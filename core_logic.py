import pandas as pd
import numpy as np
import os
import shutil
import copy
import re
import calendar
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

class DataReconciler:
    def __init__(self):
        # ตั้งค่าเริ่มต้นเป็น Shipping Keys (ถ้าระบบตรวจพบไฟล์อื่น ค่านี้จะถูกเปลี่ยนให้อัตโนมัติ)
        self.default_shipping_keys = ['Territory', 'Product Code', 'PART NO', 'Ship to']
        self.key_cols = list(self.default_shipping_keys)
        self.diff_map = {}        
        self.new_keys = set()     
        self.file2_path = ""
        self.sheet2_name = ""
        self.result_df = None

    # ─────────────────────────────────────────────
    # Public helpers
    # ─────────────────────────────────────────────
    def get_sheet_names(self, filepath):
        if filepath.lower().endswith('.xlsx'):
            try:
                return pd.ExcelFile(filepath).sheet_names
            except Exception as e:
                raise Exception(f"ไม่สามารถอ่านโครงสร้างไฟล์ Excel ได้: {str(e)}")
        return ["CSV Default"]

    def get_columns(self, filepath, sheet_name):
        return list(self.clean_and_prepare_df(filepath, sheet_name).columns)

    # ─────────────────────────────────────────────
    # Internal helpers
    # ─────────────────────────────────────────────
    def _find_header_idx(self, df):
        # 🚀 SMART TWEAK 1: ตรวจหาแบบสากล
        # ขั้นแรก: เช็คก่อนว่ามีคอลัมน์คู่ใจของระบบ Shipping เดิมไหม
        for idx, row in df.head(40).iterrows():
            vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
            if any('territory' in v for v in vals) and any('part' in v and 'no' in v.replace(' ', '') for v in vals):
                return idx
                
        # ขั้นที่สอง (Fallback): ถ้าไม่ใช่ไฟล์ Shipping ให้หาแถวที่มีตัวหนังสือหนาแน่นที่สุด (มักจะเป็นหัวตาราง)
        max_score = -1
        best_idx = 0
        for idx, row in df.head(20).iterrows():
            vals = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip() != '']
            if not vals: continue
            # ให้คะแนนแถวที่มีข้อความยาวกำลังดีและไม่ใช่ตัวเลขล้วน
            score = sum(1 for v in vals if not v.replace('.','',1).isdigit() and len(v) < 50)
            if score > max_score:
                max_score = score
                best_idx = idx
        return best_idx

    def _col_name(self, raw, i):
        if pd.isna(raw) or str(raw).strip() == '':
            return f'__blank_{i}'

        if hasattr(raw, 'strftime'):
            return raw.strftime('%d-%b')

        s = str(raw).strip()
        lo = s.lower()

        months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        if any(m in lo for m in months) and any(c.isdigit() for c in s):
            return s

        iso_match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
        if iso_match:
            y, m, d = iso_match.groups()
            try:
                return f"{int(d):02d}-{calendar.month_abbr[int(m)]}"
            except:
                pass

        # ทำความสะอาดคอลัมน์มาตรฐาน (แต่ถ้าไม่ตรง ก็จะส่งชื่อเดิมกลับไปให้)
        if 'territory' in lo:          return 'Territory'
        if 'product' in lo and 'code' in lo: return 'Product Code'
        if 'part' in lo and 'no' in lo:return 'PART NO'
        if 'ship' in lo and 'to' in lo:return 'Ship to'
        if 'carton' in lo:             return 'carton box'
        
        return s

    def _make_unique_cols(self, raw_header_row):
        seen = {}
        cols = []
        for i, raw in enumerate(raw_header_row):
            name = self._col_name(raw, i)
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 1
            cols.append(name)
        return cols

    # ─────────────────────────────────────────────
    # Core: clean & prepare dataframe
    # ─────────────────────────────────────────────
    def clean_and_prepare_df(self, filepath, sheet_name, row_range=""):
        try:
            if filepath.lower().endswith('.csv'):
                encodings = ['utf-8-sig', 'utf-8', 'cp874', 'tis-620', 'windows-1252']
                df_str = None
                for enc in encodings:
                    try:
                        df_str = pd.read_csv(filepath, header=None, encoding=enc, dtype=str, sep=None, engine='python')
                        break
                    except Exception:
                        continue
                if df_str is None or df_str.empty:
                    raise ValueError(f"ไม่พบข้อมูลใน Sheet: {sheet_name}")
                header_idx = self._find_header_idx(df_str)
                if header_idx == -1:
                    raise KeyError(f"ใน Sheet '{sheet_name}' ไม่พบหัวตาราง")
                df = df_str
                cols = self._make_unique_cols(df_str.iloc[header_idx])
            else:
                df_raw = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                if df_raw is None or df_raw.empty:
                    raise ValueError(f"ไม่พบข้อมูลใน Sheet: {sheet_name}")
                header_idx = self._find_header_idx(df_raw)
                if header_idx == -1:
                    raise KeyError(f"ใน Sheet '{sheet_name}' ไม่พบหัวตาราง")
                cols = self._make_unique_cols(df_raw.iloc[header_idx])
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=str)

            df.columns = cols
            df = df.iloc[header_idx + 1:].reset_index(drop=True)

            if row_range.strip():
                try:
                    s, e = map(int, row_range.split('-'))
                    df = df.iloc[max(0, s-1):e]
                except Exception:
                    pass

            # เคลียร์ค่าว่างเฉพาะคอลัมน์ที่เป็น Key จริงๆ ณ ขณะนั้น
            active_keys = [c for c in self.key_cols if c in df.columns]
            for col in active_keys:
                df[col] = df[col].astype(str).str.strip()
                df = df[df[col].notna() & (df[col] != '') & (df[col].str.lower() != 'nan') & (df[col].str.lower() != 'none')]

            if active_keys:
                if df.set_index(active_keys).index.duplicated().any():
                    df = df[~df.set_index(active_keys).index.duplicated(keep='first')]

            return df
        except Exception as e:
            raise Exception(f"Error in {os.path.basename(filepath)} (Sheet: {sheet_name}): {str(e)}")

    # ─────────────────────────────────────────────
    # Core: compare
    # ─────────────────────────────────────────────
    def compare_data(self, file1, sheet1, file2, sheet2, target_cols=None, row_range=""):
        self.file2_path = file2
        self.sheet2_name = sheet2
        self.diff_map = {}
        self.new_keys = set()

        # อ่านข้อมูลดิบมาก่อนเพื่อเช็คชื่อคอลัมน์
        df1 = self.clean_and_prepare_df(file1, sheet1, row_range)
        df2 = self.clean_and_prepare_df(file2, sheet2, row_range)

        # 🚀 SMART TWEAK 2: ระบบตรวจจับ Key อัตโนมัติ (Dynamic Key Detection)
        # ถ้าไฟล์มีคอลัมน์ครบตามเงื่อนไขของ Shippingเดิม ให้ใช้ของเดิม
        if all(k in df1.columns for k in self.default_shipping_keys) and all(k in df2.columns for k in self.default_shipping_keys):
            self.key_cols = list(self.default_shipping_keys)
        else:
            # ถ้าเป็นไฟล์อื่นๆ ให้มองหาคอลัมน์ที่มีคำว่า ID, Code, No, หรือ Key เป็นหลักในการจับคู่ข้อมูล
            common_cols = list(df1.columns.intersection(df2.columns))
            detected_keys = []
            for col in common_cols:
                cl = str(col).lower()
                if any(k in cl for k in ['id', 'code', 'no.', 'no ', 'key', 'part', 'serial', 'sku', 'name']):
                    detected_keys.append(col)
            
            # ถ้าหาตาม Keyword ไม่เจอเลย ให้เหมาคอลัมน์แรกของตารางเป็น Key แทน
            if not detected_keys:
                detected_keys = common_cols[:1] if common_cols else ['Index']
                
            self.key_cols = detected_keys

        # ทำการตั้ง Index ตาม Key คอลัมน์ที่ระบบวิเคราะห์ออกมาได้
        df1.set_index(self.key_cols, inplace=True)
        df2.set_index(self.key_cols, inplace=True)

        common_cols = list(df1.columns.intersection(df2.columns))
        if target_cols:
            common_cols = [c for c in common_cols if c in target_cols or c in self.key_cols]
        common_cols = pd.Index(common_cols)

        df1_c = df1[common_cols].fillna('')
        df2_c = df2[common_cols].fillna('')

        new_keys_idx = df2_c.index.difference(df1_c.index)
        df_new = df2_c.loc[new_keys_idx].copy()
        df_new['Status'] = 'New'
        df_new['Changed_Summary'] = 'รายการใหม่'
        for k in new_keys_idx:
            self.new_keys.add(tuple(str(x).strip() for x in (k if isinstance(k, tuple) else (k,))))

        del_keys_idx = df1_c.index.difference(df2_c.index)
        df_del = df1_c.loc[del_keys_idx].copy()
        df_del['Status'] = 'Deleted'
        df_del['Changed_Summary'] = 'ถูกลบทิ้ง'

        common_keys = df1_c.index.intersection(df2_c.index)
        df1_cm = df1_c.loc[common_keys].sort_index()
        df2_cm = df2_c.loc[common_keys].sort_index()
        changed_mask = (df1_cm != df2_cm).any(axis=1)

        df_changed = pd.DataFrame()
        if changed_mask.any():
            records = []
            for idx in df1_cm[changed_mask].index:
                row_data = {'Status': 'Changed'}
                if isinstance(idx, tuple):
                    row_data.update(dict(zip(self.key_cols, idx)))
                    dict_key = tuple(str(x).strip() for x in idx)
                else:
                    row_data[self.key_cols[0]] = idx
                    dict_key = (str(idx).strip(),)

                changed_cols = []
                changes = {}
                for col in common_cols:
                    v1 = str(df1_cm.loc[idx, col]).strip()
                    v2 = str(df2_cm.loc[idx, col]).strip()
                    
                    if v1.endswith('.0'): v1 = v1[:-2]
                    if v2.endswith('.0'): v2 = v2[:-2]

                    if v1 != v2:
                        val1_disp = "-" if v1 in ("", "nan") else v1
                        val2_disp = "-" if v2 in ("", "nan") else v2
                        row_data[col] = f"[ {val1_disp} ➔ {val2_disp} ]"
                        changed_cols.append(str(col))
                        changes[str(col)] = (val1_disp, val2_disp)
                    else:
                        row_data[col] = v2 if v2 != "nan" else ""

                if changed_cols: 
                    row_data['Changed_Summary'] = ', '.join(changed_cols)
                    records.append(row_data)
                    self.diff_map[dict_key] = changes

            df_changed = pd.DataFrame(records)
            if not df_changed.empty:
                df_changed.set_index(self.key_cols, inplace=True)

        result_df = pd.concat([df_changed, df_new, df_del]).reset_index()
        if not result_df.empty:
            result_df.columns = [str(c).strip() for c in result_df.columns]
            first = ['Status', 'Changed_Summary'] + self.key_cols
            others = [c for c in result_df.columns if c not in first and not c.startswith('__blank_')]
            
            # ถ้าเป็นไฟล์แบบสากล ให้แสดงผลคอลัมน์ข้อมูลทั่วไปทั้งหมดต่อท้าย Key หลัก
            result_df = result_df[first + others]

        self.result_df = result_df
        return result_df

    # ─────────────────────────────────────────────
    # Export: สร้าง Sheet "Change_Details" แบบไดนามิก
    # ─────────────────────────────────────────────
    def export_excel(self, save_path):
        if not self.file2_path.lower().endswith('.xlsx'):
            if self.result_df is not None:
                self.result_df.to_excel(save_path, index=False)
            return

        shutil.copy2(self.file2_path, save_path)
        wb = openpyxl.load_workbook(save_path)
        ws = wb[self.sheet2_name] if self.sheet2_name in wb.sheetnames else wb.active

        CHANGED_FILL = PatternFill(start_color='FFCC99FF', end_color='FFCC99FF', fill_type='solid')
        NEW_ROW_FILL  = PatternFill(start_color='FFE8CCFF', end_color='FFE8CCFF', fill_type='solid')
        
        # ค้นหาตำแหน่งหัวตารางในไฟล์ Excel ผลลัพธ์
        header_row = 1
        for r in range(1, 20):
            vals = [str(ws.cell(r, c).value or '').strip().lower() for c in range(1, 12)]
            if any(str(k).lower() in vals for k in self.key_cols):
                header_row = r
                break

        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(header_row, c).value
            if val is None: continue
            name = self._col_name(val, c)
            col_map[name] = c

        change_details_data = []

        for r in range(header_row + 1, ws.max_row + 1):
            row_key = []
            for k in self.key_cols:
                if k in col_map:
                    v = ws.cell(r, col_map[k]).value
                    row_key.append(str(v).strip() if v is not None else 'nan')
                else:
                    row_key.append('nan')
            row_key_tuple = tuple(row_key)

            carton_val = ""
            if 'carton box' in col_map:
                cv = ws.cell(r, col_map['carton box']).value
                carton_val = str(cv).strip() if cv is not None and str(cv).strip() not in ('nan', 'None') else ""

            # ข้อมูลแถวที่มีการเปลี่ยนแปลง
            if row_key_tuple in self.diff_map:
                changes = self.diff_map[row_key_tuple]
                for date_col_name, (old_v, new_v) in changes.items():
                    if date_col_name in col_map:
                        c_idx = col_map[date_col_name]
                        cell = ws.cell(r, c_idx)
                        
                        orig_fill = None
                        if cell.fill and cell.fill.fill_type == 'solid':
                            orig_fill = copy.copy(cell.fill)

                        cell.fill = CHANGED_FILL
                        
                        # สร้าง Dictionary เก็บข้อมูลแบบ Dynamic ตามคอลัมน์ Key จริง
                        item_entry = {k: v for k, v in zip(self.key_cols, row_key)}
                        if 'carton box' in col_map:
                            item_entry['carton box'] = carton_val
                        item_entry.update({
                            'Date': date_col_name,
                            'Old Qty': old_v,
                            'New Qty': new_v,
                            'Orig_Fill': orig_fill
                        })
                        change_details_data.append(item_entry)

            # ข้อมูลแถวที่เพิ่มเข้ามาใหม่
            elif row_key_tuple in self.new_keys:
                max_col = ws.max_column
                for c_idx in range(1, max_col + 1):
                    ws.cell(r, c_idx).fill = NEW_ROW_FILL
                
                for col_name, c_idx in col_map.items():
                    if col_name in self.key_cols or col_name == 'carton box': continue
                    cell = ws.cell(r, c_idx)
                    
                    if cell.value is not None and str(cell.value).strip() not in ('', 'nan', '0', 'None'):
                        orig_fill = None
                        if cell.fill and cell.fill.fill_type == 'solid':
                            orig_fill = copy.copy(cell.fill)

                        item_entry = {k: v for k, v in zip(self.key_cols, row_key)}
                        if 'carton box' in col_map:
                            item_entry['carton box'] = carton_val
                        item_entry.update({
                            'Date': col_name,
                            'Old Qty': '-',
                            'New Qty': str(cell.value),
                            'Orig_Fill': orig_fill
                        })
                        change_details_data.append(item_entry)

        # สร้าง Sheet ใหม่ชื่อ Change_Details
        if 'Change_Details' in wb.sheetnames: del wb['Change_Details']
        if 'Transport_Report' in wb.sheetnames: del wb['Transport_Report']
            
        ws_det = wb.create_sheet('Change_Details', 0)
        
        # 🚀 SMART TWEAK 3: ประกอบหัวข้อรายงานตามโครงสร้างจริงของไฟล์นั้นๆ
        headers = list(self.key_cols)
        if 'carton box' in col_map:
            headers.append('carton box')
        headers.extend(['Date', 'Old Qty', 'New Qty'])
        
        ws_det.append(headers)
        
        HDR_FILL = PatternFill(start_color='FF1E293B', end_color='FF1E293B', fill_type='solid')
        HDR_FONT = Font(bold=True, color='FFFFFF')
        for c in range(1, len(headers) + 1):
            cell = ws_det.cell(1, c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT

        # บันทึกข้อมูลลงตาราง
        for item in change_details_data:
            old_v = str(item['Old Qty']).strip()
            new_v = str(item['New Qty']).strip()
            
            old_text = "ค่าเดิมไม่มี" if old_v in ('-', '', 'nan', '0', 'None') else f"ค่าเดิม {old_v}"
            new_text = "ค่าใหม่ไม่มี" if new_v in ('-', '', 'nan', '0', 'None') else f"ค่าใหม่ {new_v}"
            
            row_vals = []
            for h in headers:
                if h == 'Date': row_vals.append(item['Date'])
                elif h == 'Old Qty': row_vals.append(old_text)
                elif h == 'New Qty': row_vals.append(new_text)
                else: row_vals.append(item.get(h, '').replace('nan', ''))
                
            ws_det.append(row_vals)
            current_row = ws_det.max_row
            
            # คืนค่าสีสันไฮไลท์ให้ช่อง Date
            if item['Orig_Fill']:
                # หาตำแหน่งหลักของ 'Date' เพื่อหยอดสีให้ถูกช่อง
                date_col_idx = headers.index('Date') + 1
                date_cell = ws_det.cell(current_row, date_col_idx)
                date_cell.fill = item['Orig_Fill']
                date_cell.font = Font(color='000000')

        # ขยายความกว้างของทุกคอลัมน์อัตโนมัติเพื่อไม่ให้โดนบัง
        for col in ws_det.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_det.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(save_path)